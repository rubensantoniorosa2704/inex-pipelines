"""
scripts/inspect_igc.py — Inspeciona todos os arquivos bronze do IGC.

Para cada ano, lista:
  - Nome do arquivo
  - Nomes das colunas (header)
  - Primeiras linhas de dados (para inferir tipos)
  - Quantidade de linhas

Uso:
  python scripts/inspect_igc.py
"""

from pathlib import Path

import openpyxl
import xlrd


BRONZE_IGC = Path("data/bronze/igc")


def read_xls_info(path: Path) -> dict:
    """Lê info de arquivo .xls via xlrd."""
    wb = xlrd.open_workbook(str(path))
    ws = wb.sheet_by_index(0)

    header = [str(ws.cell_value(0, c)).strip() for c in range(ws.ncols)]
    nrows = ws.nrows - 1  # exclui header

    # Primeiras 3 linhas para amostra
    sample_rows = []
    for r in range(1, min(4, ws.nrows)):
        row = []
        for c in range(ws.ncols):
            cell = ws.cell(r, c)
            if cell.ctype == 0:
                row.append(None)
            elif cell.ctype == 2:
                val = cell.value
                row.append(int(val) if val == int(val) else val)
            else:
                row.append(str(cell.value).strip() if cell.value != "" else None)
        sample_rows.append(row)

    return {"header": header, "nrows": nrows, "sample": sample_rows, "sheets": [wb.sheet_by_index(i).name for i in range(wb.nsheets)]}


def read_xlsx_info(path: Path) -> dict:
    """Lê info de arquivo .xlsx via openpyxl."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    # Pega a primeira sheet relevante
    skip_sheets = {"Atualizações", "Plan2", "Plan3"}
    sheet_name = next((s for s in wb.sheetnames if s not in skip_sheets), wb.sheetnames[0])
    ws = wb[sheet_name]

    rows = list(ws.iter_rows(min_row=1, max_row=5, values_only=True))
    wb.close()

    if not rows:
        return {"header": [], "nrows": 0, "sample": [], "sheets": wb.sheetnames}

    header = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]

    # Para contar linhas, reabrir (read_only não tem max_row confiável)
    wb2 = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws2 = wb2[sheet_name]
    nrows = sum(1 for _ in ws2.iter_rows(min_row=2, values_only=True))
    wb2.close()

    sample_rows = []
    for row in rows[1:]:
        sample_rows.append([
            str(v).strip() if v is not None else None
            for v in row
        ])

    return {"header": header, "nrows": nrows, "sample": sample_rows, "sheets": list(wb.sheetnames) if hasattr(wb, 'sheetnames') else []}


def inspect_all():
    years = sorted([d.name for d in BRONZE_IGC.iterdir() if d.is_dir()])

    all_columns = {}  # coluna → set de anos onde aparece

    for year_str in years:
        year_dir = BRONZE_IGC / year_str
        files = list(year_dir.glob("*.*"))
        if not files:
            print(f"\n{'='*60}\n{year_str}: VAZIO\n{'='*60}")
            continue

        f = files[0]
        print(f"\n{'='*60}")
        print(f"ANO: {year_str} | Arquivo: {f.name} | Tamanho: {f.stat().st_size / 1024:.0f} KB")
        print(f"{'='*60}")

        try:
            if f.suffix.lower() == ".xls":
                info = read_xls_info(f)
            else:
                info = read_xlsx_info(f)

            print(f"  Sheets: {info.get('sheets', ['?'])}")
            print(f"  Linhas de dados: {info['nrows']}")
            print(f"  Colunas ({len(info['header'])}):")
            for i, col in enumerate(info['header']):
                sample_vals = [row[i] for row in info['sample'] if i < len(row)]
                sample_str = " | ".join(str(v) for v in sample_vals[:3])
                print(f"    [{i:2d}] {col!r:50s} ex: {sample_str}")

                # Rastrear em quais anos cada coluna aparece
                if col not in all_columns:
                    all_columns[col] = []
                all_columns[col].append(year_str)

        except Exception as e:
            print(f"  ERRO: {e}")

    # Resumo: todas as colunas encontradas em todos os anos
    print(f"\n\n{'='*60}")
    print("RESUMO: TODAS AS COLUNAS ENCONTRADAS")
    print(f"{'='*60}")
    for col, col_years in sorted(all_columns.items(), key=lambda x: (-len(x[1]), x[0])):
        print(f"  {col!r:50s} → {len(col_years)} anos: {col_years}")


if __name__ == "__main__":
    inspect_all()
