"""
shared/lookup_cocurso.py — Lookup de co_curso via microdados do ENADE e conceito-enade.

Fornece funções para preencher co_curso nas tabelas do CPC (2007–2016)
onde o INEP não publicava essa informação.

Estratégia em cascata:
  1. Conceito ENADE direto (2015–2016): fonte autoritativa, co_curso já existe
  2. Microdados ENADE — nível 1: (co_ies, co_grupo) unívoco → co_curso
  3. Microdados ENADE — nível 2: (co_ies, co_grupo, co_munic_curso) → co_curso

Cobertura medida: ~78.5% dos registros 2007–2016.
"""

from pathlib import Path

import polars as pl

from shared.paths import BRONZE_ROOT


# ---------------------------------------------------------------------------
# Localização de arquivos
# ---------------------------------------------------------------------------

def _find_enade_arq1(year: int) -> Path | None:
    """Localiza o arq1.txt dos microdados do ENADE para um ano."""
    year_dir = BRONZE_ROOT / "enade-microdados" / str(year)
    if not year_dir.exists():
        return None
    candidates = list(year_dir.rglob(f"microdados{year}_arq1.txt"))
    if not candidates:
        candidates = list(year_dir.rglob("*arq1.txt"))
    return candidates[0] if candidates else None


def _find_conceito_enade(year: int) -> Path | None:
    """Localiza o arquivo do conceito-enade para um ano."""
    year_dir = BRONZE_ROOT / "conceito-enade" / str(year)
    if not year_dir.exists():
        return None
    files = (
        list(year_dir.glob("*.xlsx")) + list(year_dir.glob("*.XLSX")) +
        list(year_dir.glob("*.xls")) + list(year_dir.glob("*.XLS"))
    )
    return files[0] if files else None


# ---------------------------------------------------------------------------
# Leitura dos microdados ENADE
# ---------------------------------------------------------------------------

def _read_enade_lookup(path: Path) -> pl.DataFrame:
    """
    Lê colunas-chave do arq1.txt do ENADE e retorna combinações distintas de
    (co_ies, co_grupo, co_curso, co_munic_curso).
    """
    # Detecta encoding
    encoding = "utf-8"
    try:
        pl.read_csv(path, separator=";", encoding="utf-8", n_rows=0)
    except Exception:
        encoding = "latin-1"

    # Lê header para identificar colunas
    df_header = pl.read_csv(path, separator=";", encoding=encoding, n_rows=0)

    col_map: dict[str, str] = {}
    for c in df_header.columns:
        cu = c.strip('"').upper()
        if cu == "CO_IES":
            col_map["co_ies"] = c
        elif cu == "CO_GRUPO":
            col_map["co_grupo"] = c
        elif cu == "CO_CURSO":
            col_map["co_curso"] = c
        elif cu == "CO_MUNIC_CURSO":
            col_map["co_munic_curso"] = c

    required = ["co_ies", "co_grupo", "co_curso"]
    if not all(k in col_map for k in required):
        raise ValueError(
            f"Colunas necessárias não encontradas em {path.name}. "
            f"Disponíveis: {df_header.columns}"
        )

    cols_to_read = list(col_map.values())
    df = pl.read_csv(
        path,
        separator=";",
        encoding=encoding,
        columns=cols_to_read,
        schema_overrides={c: pl.Utf8 for c in cols_to_read},
    )

    # Renomeia para nomes canônicos
    rename = {v: k for k, v in col_map.items()}
    df = df.rename(rename)

    # Cast para Int64
    for col in df.columns:
        df = df.with_columns(
            pl.col(col).str.strip_chars('" ').cast(pl.Int64, strict=False).alias(col)
        )

    return df.unique().drop_nulls()


# ---------------------------------------------------------------------------
# Leitura do conceito-enade (2015–2016 têm co_curso)
# ---------------------------------------------------------------------------

CONCEITO_ENADE_YEARS_WITH_CURSO = {2015, 2016}

# Variações de nomes de colunas no conceito-enade
_CONCEITO_COL_VARIANTS = {
    "co_ies": ["Código da IES", "Código IES", "co_ies"],
    "co_curso": ["Código do Curso", "Código do Curso**"],
    "co_area": ["Código da Área", "Código Área", "Cód.Área"],
    "co_municipio": ["Código do Município", "Código do Município***", "Cód. Município"],
}


def _read_conceito_enade_lookup(path: Path) -> pl.DataFrame | None:
    """
    Lê o conceito-enade e extrai (co_ies, co_area, co_curso, co_municipio).
    Retorna None se o arquivo não tiver co_curso.
    """
    suffix = path.suffix.lower()

    if suffix == ".xls":
        import xlrd
        wb = xlrd.open_workbook(str(path))
        ws = wb.sheet_by_index(0)
        header = [str(ws.cell_value(0, c)).strip() for c in range(ws.ncols)]
        rows = []
        for r in range(1, ws.nrows):
            rows.append([str(ws.cell_value(r, c)).strip() for c in range(ws.ncols)])
        data = {col: [row[i] for row in rows] for i, col in enumerate(header)}
        df = pl.DataFrame(data)
    else:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        skip = {"Atualizações", "Plan2", "Plan3", "Legenda", "Notas"}
        sheet_name = next((s for s in wb.sheetnames if s not in skip), wb.sheetnames[0])
        ws = wb[sheet_name]
        rows_raw = list(ws.iter_rows(values_only=True))
        if not rows_raw:
            return None
        header = [str(h).strip() if h is not None else f"col_{i}"
                  for i, h in enumerate(rows_raw[0])]
        data = {
            col: [str(rows_raw[r][i]).strip() if rows_raw[r][i] is not None else None
                  for r in range(1, len(rows_raw))]
            for i, col in enumerate(header)
        }
        df = pl.DataFrame(data)

    # Identifica colunas
    found: dict[str, str] = {}
    for canonical, variants in _CONCEITO_COL_VARIANTS.items():
        for v in variants:
            if v in df.columns:
                found[canonical] = v
                break

    if "co_curso" not in found or "co_ies" not in found:
        return None

    # Seleciona e renomeia
    select_cols = {v: k for k, v in found.items()}
    df = df.select(list(select_cols.keys())).rename(select_cols)

    # Cast para Int64
    for col in df.columns:
        df = df.with_columns(
            pl.col(col).str.strip_chars('" ').cast(pl.Int64, strict=False).alias(col)
        )

    # Filtra nulos e linhas inválidas
    df = df.filter(pl.col("co_ies").is_not_null() & pl.col("co_curso").is_not_null())
    return df.unique()


# ---------------------------------------------------------------------------
# API pública: enriquecer DataFrame do CPC com co_curso
# ---------------------------------------------------------------------------

def enrich_cocurso(cpc_df: pl.DataFrame, year: int, verbose: bool = False) -> pl.DataFrame:
    """
    Tenta preencher co_curso em um DataFrame do CPC silver para um ano específico.

    Cascata:
      1. Conceito ENADE direto (se disponível para o ano)
      2. Microdados ENADE — nível 1: (co_ies, co_area) unívoco
      3. Microdados ENADE — nível 2: (co_ies, co_area, co_municipio) unívoco

    Retorna o DataFrame com co_curso preenchido onde possível.
    Linhas sem match mantêm co_curso nulo.
    """
    if "co_area" not in cpc_df.columns:
        if verbose:
            print(f"    [lookup] co_area não existe no CPC {year}, pulando")
        return cpc_df

    # Garante que co_curso existe como coluna (pode não existir nos antigos)
    if "co_curso" not in cpc_df.columns:
        cpc_df = cpc_df.with_columns(pl.lit(None).cast(pl.Int64).alias("co_curso"))

    total = cpc_df.shape[0]
    preenchidos = 0

    # --- Estratégia 1: Conceito ENADE (2015–2016) ---
    if year in CONCEITO_ENADE_YEARS_WITH_CURSO:
        conceito_path = _find_conceito_enade(year)
        if conceito_path:
            conceito = _read_conceito_enade_lookup(conceito_path)
            if conceito is not None:
                # Join direto por (co_ies, co_area)
                join_cols = ["co_ies", "co_area"]
                if "co_municipio" in conceito.columns and "co_municipio" in cpc_df.columns:
                    join_cols.append("co_municipio")

                lookup = conceito.select(
                    [pl.col(c) for c in join_cols] + [pl.col("co_curso").alias("_co_curso_conceito")]
                ).unique()

                # Desambigua: só usa pares unívocos
                key_cols = [c for c in join_cols]
                count = lookup.group_by(key_cols).agg(pl.len().alias("n"))
                univocos = count.filter(pl.col("n") == 1).select(key_cols)
                lookup = lookup.join(univocos, on=key_cols, how="inner")

                cpc_df = cpc_df.join(lookup, on=key_cols, how="left")
                cpc_df = cpc_df.with_columns(
                    pl.coalesce(["co_curso", "_co_curso_conceito"]).alias("co_curso")
                ).drop("_co_curso_conceito")

                preenchidos = cpc_df.filter(pl.col("co_curso").is_not_null()).shape[0]
                if verbose:
                    print(f"    [conceito-enade] {preenchidos}/{total} preenchidos")

                if preenchidos == total:
                    return cpc_df

    # --- Estratégias 2 e 3: Microdados ENADE ---
    arq1_path = _find_enade_arq1(year)
    if arq1_path is None:
        if verbose:
            print(f"    [lookup] Microdados ENADE {year} não encontrados")
        return cpc_df

    try:
        enade = _read_enade_lookup(arq1_path)
    except Exception as e:
        if verbose:
            print(f"    [lookup] Erro lendo ENADE {year}: {e}")
        return cpc_df

    has_munic = "co_munic_curso" in enade.columns

    # --- Nível 1: (co_ies, co_grupo) unívoco ---
    ainda_nulos = cpc_df.filter(pl.col("co_curso").is_null())
    if ainda_nulos.shape[0] > 0:
        cursos_por_par = (
            enade.select(["co_ies", "co_grupo", "co_curso"]).unique()
            .group_by(["co_ies", "co_grupo"])
            .agg(
                pl.col("co_curso").n_unique().alias("n_cursos"),
                pl.col("co_curso").first().alias("co_curso_unico"),
            )
        )
        univocos_n1 = cursos_por_par.filter(pl.col("n_cursos") == 1).select(
            pl.col("co_ies"),
            pl.col("co_grupo").alias("co_area"),
            pl.col("co_curso_unico").alias("_co_curso_n1"),
        )

        cpc_df = cpc_df.join(univocos_n1, on=["co_ies", "co_area"], how="left")
        cpc_df = cpc_df.with_columns(
            pl.coalesce(["co_curso", "_co_curso_n1"]).alias("co_curso")
        ).drop("_co_curso_n1")

        n1_preenchidos = cpc_df.filter(pl.col("co_curso").is_not_null()).shape[0]
        if verbose:
            print(f"    [enade-n1] {n1_preenchidos}/{total} preenchidos (+{n1_preenchidos - preenchidos})")
        preenchidos = n1_preenchidos

    # --- Nível 2: (co_ies, co_grupo, co_munic_curso) unívoco ---
    if has_munic and "co_municipio" in cpc_df.columns:
        ainda_nulos = cpc_df.filter(pl.col("co_curso").is_null())
        if ainda_nulos.shape[0] > 0:
            cursos_por_trio = (
                enade.select(["co_ies", "co_grupo", "co_munic_curso", "co_curso"]).unique()
                .group_by(["co_ies", "co_grupo", "co_munic_curso"])
                .agg(
                    pl.col("co_curso").n_unique().alias("n_cursos"),
                    pl.col("co_curso").first().alias("co_curso_unico"),
                )
            )
            univocos_n2 = cursos_por_trio.filter(pl.col("n_cursos") == 1).select(
                pl.col("co_ies"),
                pl.col("co_grupo").alias("co_area"),
                pl.col("co_munic_curso").alias("co_municipio"),
                pl.col("co_curso_unico").alias("_co_curso_n2"),
            )

            cpc_df = cpc_df.join(
                univocos_n2, on=["co_ies", "co_area", "co_municipio"], how="left"
            )
            cpc_df = cpc_df.with_columns(
                pl.coalesce(["co_curso", "_co_curso_n2"]).alias("co_curso")
            ).drop("_co_curso_n2")

            n2_preenchidos = cpc_df.filter(pl.col("co_curso").is_not_null()).shape[0]
            if verbose:
                print(f"    [enade-n2] {n2_preenchidos}/{total} preenchidos (+{n2_preenchidos - preenchidos})")
            preenchidos = n2_preenchidos

    if verbose:
        restantes = total - preenchidos
        print(f"    [total] {preenchidos}/{total} ({preenchidos/total*100:.1f}%), restantes: {restantes}")

    return cpc_df
