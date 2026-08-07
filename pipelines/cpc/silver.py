"""
pipelines/cpc/silver.py — Silver do CPC (Conceito Preliminar de Curso).

Lê os XLSX/XLS do INEP e produz um Parquet silver por ano:
  silver/cpc/{ano}.parquet

Uso:
  python -m pipelines.cpc.silver --year 2023
  python -m pipelines.cpc.silver --year 2007-2023
"""

import sys
from pathlib import Path

import click
import polars as pl

from pipelines.cpc.schema import (
    CPC_COLUMN_MAP,
    CPC_REQUIRED_COLUMNS,
    CPC_SILVER_SCHEMA,
    YEARS_WITHOUT_CO_CURSO,
    YEARS_WITHOUT_CPC_FAIXA,
)
from shared.io import write_parquet
from shared.paths import bronze_dir, silver_path
from shared.validate import assert_no_nulls, assert_not_empty, assert_required_columns


def _find_file(year: int) -> Path:
    """Localiza o arquivo Excel (xlsx ou xls) do CPC para o ano dado."""
    base = bronze_dir("cpc", year)
    if not base.exists():
        raise FileNotFoundError(
            f"Diretório bronze não encontrado: {base}\n"
            f"Execute: inex-ingest ingest --dataset cpc --year {year}"
        )
    candidates = (
        list(base.glob("*.xlsx")) + list(base.glob("*.XLSX")) +
        list(base.glob("*.xls"))  + list(base.glob("*.XLS"))
    )
    if not candidates:
        raise FileNotFoundError(f"Nenhum arquivo Excel encontrado em {base}")
    # Prefere xlsx sobre xls quando ambos existem
    xlsx = [f for f in candidates if f.suffix.lower() == ".xlsx"]
    return xlsx[0] if xlsx else candidates[0]


def _read_xls(path: Path) -> pl.DataFrame:
    """Lê arquivos .xls antigos via xlrd."""
    import xlrd

    wb = xlrd.open_workbook(str(path))
    ws = wb.sheet_by_index(0)

    header = [str(ws.cell_value(0, c)).strip() for c in range(ws.ncols)]
    rows = []
    for r in range(1, ws.nrows):
        row = []
        for c in range(ws.ncols):
            cell = ws.cell(r, c)
            # xlrd tipo 0=empty, 1=text, 2=number, 3=date, 5=bool
            if cell.ctype == 0:
                row.append(None)
            elif cell.ctype == 2:
                # Números: preserva inteiros como inteiros
                val = cell.value
                row.append(str(int(val)) if val == int(val) else str(val))
            else:
                row.append(str(cell.value).strip() if cell.value != "" else None)
        rows.append(row)

    str_data = {col: [row[i] if i < len(row) else None for row in rows]
                for i, col in enumerate(header)}
    df = pl.DataFrame(str_data)

    # Remove linhas de rodapé onde o código da IES não é numérico
    ies_col = next((c for c in ["co_ies", "Código da IES", "Código da IES*"]
                    if c in df.columns), None)
    if ies_col:
        df = df.filter(pl.col(ies_col).str.contains(r"^\d+$"))

    return df.filter(pl.any_horizontal(pl.all().is_not_null()))


def _read_xlsx(path: Path) -> pl.DataFrame:
    """Lê arquivos .xlsx via openpyxl."""
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    skip_sheets = {"Atualizações", "Plan2", "Plan3"}
    sheet_name = next((s for s in wb.sheetnames if s not in skip_sheets), wb.sheetnames[0])
    ws = wb[sheet_name]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"Planilha vazia: {path}")

    header = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
    data = rows[1:]

    str_data = {
        col: [
            str(row[i]).strip() if i < len(row) and row[i] is not None else None
            for row in data
        ]
        for i, col in enumerate(header)
    }

    df = pl.DataFrame(str_data)

    ies_col = next(
        (c for c in ["Código da IES", "Código da IES*", "co_ies"] if c in df.columns),
        None,
    )
    if ies_col:
        df = df.filter(pl.col(ies_col).str.contains(r"^\d+$"))

    return df.filter(pl.any_horizontal(pl.all().is_not_null()))


def _read_file(path: Path) -> pl.DataFrame:
    """Despacha para o leitor correto com base na extensão."""
    if path.suffix.lower() == ".xls":
        return _read_xls(path)
    return _read_xlsx(path)


def _rename_columns(df: pl.DataFrame) -> pl.DataFrame:
    """
    Renomeia colunas usando o mapa canônico.
    - Descarta colunas auxiliares (prefixo _)
    - Quando duas colunas originais mapeiam para o mesmo nome canônico,
      mantém a primeira ocorrência e descarta as demais (evita duplicatas).
    """
    seen_targets: set[str] = set()
    rename_map: dict[str, str] = {}
    drop_cols: list[str] = []

    for orig_col in df.columns:
        target = CPC_COLUMN_MAP.get(orig_col)
        if target is None:
            continue
        if target.startswith("_"):
            drop_cols.append(orig_col)
        elif target in seen_targets:
            # Segunda ocorrência do mesmo destino — descarta
            drop_cols.append(orig_col)
        else:
            rename_map[orig_col] = target
            seen_targets.add(target)

    df = df.drop([c for c in drop_cols if c in df.columns])
    return df.rename(rename_map)


def _cast_schema(df: pl.DataFrame) -> pl.DataFrame:
    """Aplica os tipos canônicos às colunas presentes."""
    casts = [
        pl.col(col).cast(dtype, strict=False)
        for col, dtype in CPC_SILVER_SCHEMA.items()
        if col in df.columns
    ]
    return df.with_columns(casts) if casts else df


def _is_stale(year: int) -> bool:
    """Retorna True se o silver precisa ser reprocessado."""
    out = silver_path("cpc", year)
    if not out.exists():
        return True
    src = _find_file(year)
    return src.stat().st_mtime > out.stat().st_mtime


def process_year(year: int, verbose: bool = False, force: bool = False) -> None:
    """Processa um único ano do CPC."""
    file_path = _find_file(year)

    if not force and not _is_stale(year):
        if verbose:
            print(f"[{year}] silver já atualizado, pulando (use --force para reprocessar)")
        else:
            print(f"~ {year} (já atualizado)")
        return

    if verbose:
        print(f"[{year}] lendo {file_path.name}")

    raw = _read_file(file_path)

    if verbose:
        print(f"[{year}] {raw.shape[0]} linhas, colunas: {raw.columns}")

    df = raw.pipe(_rename_columns).pipe(_cast_schema)

    # Preenche 'ano' quando omitido em linhas subsequentes
    if "ano" in df.columns and df["ano"].null_count() > 0:
        first_ano = df["ano"].drop_nulls().first()
        df = df.with_columns(pl.col("ano").fill_null(first_ano))

    # Anos sem 'ano' no arquivo: injeta o ano do nome do diretório
    if "ano" not in df.columns:
        df = df.with_columns(pl.lit(year).cast(pl.Int32).alias("ano"))

    # Anos sem co_curso: não é erro — o INEP não publicava o código antes de 2017
    # Anos sem cpc_faixa: 2007/2008 só publicaram o CPC contínuo
    base_required = [c for c in CPC_REQUIRED_COLUMNS
                     if not (c == "cpc_faixa" and year in YEARS_WITHOUT_CPC_FAIXA)]
    required = (
        base_required
        if year in YEARS_WITHOUT_CO_CURSO
        else base_required + ["co_curso"]
    )

    assert_not_empty(df, f"cpc/{year}")
    assert_required_columns(df, required, f"cpc/{year}")
    assert_no_nulls(df, ["co_ies", "ano"], f"cpc/{year}")

    out = silver_path("cpc", year)
    write_parquet(df, out)

    n_cursos = df.shape[0]
    has_curso = "co_curso" in df.columns
    print(f"✓ {year} ({n_cursos} cursos{'' if has_curso else ', sem co_curso'})")
    if verbose:
        print(df.head(3))


@click.command()
@click.option("--year", required=True, help="Ano ou intervalo: 2023 ou 2007-2023")
@click.option("--force", is_flag=True, help="Reprocessa mesmo se silver já atualizado")
@click.option("--verbose", is_flag=True, help="Log detalhado")
def main(year: str, force: bool, verbose: bool) -> None:
    """Gera o silver do CPC (suporta 2007–2023, exceto 2020)."""
    from pipelines.censo.silver import _parse_years

    years = _parse_years(year)
    errors = []

    for y in years:
        if y == 2020:
            print(f"~ 2020 (não existe — ENADE não aplicado na pandemia)")
            continue
        try:
            process_year(y, verbose=verbose, force=force)
        except FileNotFoundError as e:
            print(f"✗ {y}: {e}", file=sys.stderr)
            errors.append(y)
        except Exception as e:
            print(f"✗ {y}: erro inesperado — {e}", file=sys.stderr)
            if verbose:
                import traceback
                traceback.print_exc()
            errors.append(y)

    if errors:
        print(f"\nFalhou em {len(errors)} ano(s): {errors}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
