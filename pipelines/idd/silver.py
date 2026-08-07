"""
pipelines/idd/silver.py — Silver do IDD.

Lê os XLSX do INEP e produz um Parquet silver por ano:
  silver/idd/{ano}.parquet

Uso:
  python -m pipelines.idd.silver --year 2023
  python -m pipelines.idd.silver --year 2016-2023
"""

import sys
from pathlib import Path

import click
import polars as pl

from pipelines.idd.schema import (
    IDD_COLUMN_MAP,
    IDD_FAIXA_SC,
    IDD_REQUIRED_COLUMNS,
    IDD_SILVER_SCHEMA,
)
from shared.io import write_parquet
from shared.paths import bronze_dir, silver_path
from shared.validate import assert_no_nulls, assert_not_empty, assert_required_columns


def _find_xlsx(year: int) -> Path:
    """Localiza o arquivo XLSX do IDD para o ano dado."""
    base = bronze_dir("idd", year)
    if not base.exists():
        raise FileNotFoundError(
            f"Diretório bronze não encontrado: {base}\n"
            f"Execute: inex-ingest ingest --dataset idd --year {year}"
        )
    candidates = list(base.glob("*.xlsx")) + list(base.glob("*.XLSX"))
    if not candidates:
        raise FileNotFoundError(f"Nenhum XLSX encontrado em {base}")
    return candidates[0]


def _read_xlsx(path: Path) -> pl.DataFrame:
    """
    Lê o XLSX do IDD.
    - Detecta a aba correta (ignora 'Atualizações', 'Plan2', 'Plan3')
    - Remove espaços em branco dos nomes de colunas
    - Converte todas as células para string para evitar tipos mistos
    - Descarta linhas de rodapé/nota (onde 'Código da IES' não é numérico)
    """
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    # Escolhe a aba principal — ignora abas auxiliares conhecidas
    skip_sheets = {"Atualizações", "Plan2", "Plan3"}
    sheet_name = next((s for s in wb.sheetnames if s not in skip_sheets), wb.sheetnames[0])
    ws = wb[sheet_name]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"Planilha vazia: {path}")

    # Header na primeira linha — strip de espaços
    header = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
    data = rows[1:]

    # Converte tudo para string — evita tipos mistos entre células da mesma coluna
    # O cast para os tipos corretos é feito depois no _cast_schema
    str_data = {
        col: [str(row[i]).strip() if i < len(row) and row[i] is not None else None for row in data]
        for i, col in enumerate(header)
    }

    df = pl.DataFrame(str_data)

    # Descarta linhas de rodapé: onde a coluna de IES não é numérica
    ies_col = next((c for c in ["Código da IES", "co_ies"] if c in df.columns), None)
    if ies_col:
        df = df.filter(pl.col(ies_col).str.contains(r"^\d+$"))

    return df.filter(pl.any_horizontal(pl.all().is_not_null()))


def _rename_columns(df: pl.DataFrame) -> pl.DataFrame:
    """Renomeia colunas presentes usando o mapa canônico."""
    existing = {k: v for k, v in IDD_COLUMN_MAP.items() if k in df.columns}
    return df.rename(existing)


def _normalize_cebas(df: pl.DataFrame) -> pl.DataFrame:
    """
    Converte in_cebas de string para Boolean.
    No XLSX do INEP: "X" = é CEBAS, "-" = não é CEBAS.
    Qualquer outro valor (ou ausência da coluna) fica como null.
    """
    if "in_cebas" not in df.columns:
        return df
    return df.with_columns(
        pl.when(pl.col("in_cebas") == "X")
        .then(pl.lit(True))
        .when(pl.col("in_cebas") == "-")
        .then(pl.lit(False))
        .otherwise(pl.lit(None, dtype=pl.Boolean))
        .alias("in_cebas")
    )


def _normalize_sem_conceito(df: pl.DataFrame) -> pl.DataFrame:
    """
    Deriva a flag in_sem_conceito a partir do valor bruto da coluna idd_faixa
    e, quando presente, da coluna observacao.

    O INEP sinaliza cursos sem conceito de duas formas dependendo do ano:
      - idd_faixa = 'SC'  (2016, 2017, 2019, 2022, 2023)
      - idd_faixa vazia + observacao preenchida (2021: "Curso para o qual
        estatisticamente não foi possível calcular o indicador")

    Após o cast para Int32, 'SC' vira null — sem essa flag não seria possível
    distinguir um curso SC de um curso simplesmente ausente.

    A detecção ocorre antes do cast de schema: zeramos idd_faixa/
    idd_continuo/nota_bruta_idd para null explícito nos cursos SC.
    """
    if "idd_faixa" not in df.columns:
        return df.with_columns(pl.lit(False).alias("in_sem_conceito"))

    faixa_str = pl.col("idd_faixa").cast(pl.Utf8).str.strip_chars()
    is_sc_faixa = faixa_str == IDD_FAIXA_SC

    # 2021: faixa vazia mas observacao preenchida indica o mesmo estado SC
    if "observacao" in df.columns:
        obs_str = pl.col("observacao").cast(pl.Utf8).str.strip_chars()
        faixa_vazia = faixa_str.is_null() | (faixa_str == "") | (faixa_str == "None")
        is_sc_obs = faixa_vazia & obs_str.is_not_null() & (obs_str != "") & (obs_str != "None")
        is_sc = is_sc_faixa | is_sc_obs
    else:
        is_sc = is_sc_faixa

    df = df.with_columns(is_sc.alias("in_sem_conceito"))

    # Garante que cursos SC têm todas as métricas nulas
    for col in ("idd_faixa", "idd_continuo", "nota_bruta_idd"):
        if col in df.columns:
            df = df.with_columns(
                pl.when(pl.col("in_sem_conceito"))
                .then(pl.lit(None))
                .otherwise(pl.col(col))
                .alias(col)
            )

    return df


def _fix_outlier_continuo(df: pl.DataFrame) -> pl.DataFrame:
    """
    Corrige idd_continuo para outliers negativos extremos.

    Em 2018, o INEP deixou IDD (Contínuo) vazio para cursos cujo ZIDD foi
    inferior a -3 (outliers negativos). Pela metodologia (Nota Técnica nº
    7/2024, §8.9), cursos com ZIDD < -3 recebem nota padronizada 0,
    equivalente a idd_faixa=1. O pipeline preenche idd_continuo=0.0 nesses
    casos para preservar a semântica da escala 0–5.

    Condição: nota_bruta_idd preenchida + idd_continuo null + idd_faixa=1
              + in_sem_conceito=False.
    """
    if "idd_continuo" not in df.columns or "idd_faixa" not in df.columns:
        return df

    sc_flag = pl.col("in_sem_conceito") if "in_sem_conceito" in df.columns else pl.lit(False)

    is_outlier_neg = (
        pl.col("nota_bruta_idd").is_not_null()
        & pl.col("idd_continuo").is_null()
        & (pl.col("idd_faixa") == 1)
        & sc_flag.not_()
    )

    return df.with_columns(
        pl.when(is_outlier_neg)
        .then(pl.lit(0.0))
        .otherwise(pl.col("idd_continuo"))
        .alias("idd_continuo")
    )


def _cast_schema(df: pl.DataFrame) -> pl.DataFrame:
    """Aplica os tipos canônicos às colunas presentes."""
    casts = [
        pl.col(col).cast(dtype, strict=False)
        for col, dtype in IDD_SILVER_SCHEMA.items()
        if col in df.columns
    ]
    return df.with_columns(casts) if casts else df


def _is_stale(year: int) -> bool:
    """Verifica se o silver precisa ser reprocessado."""
    out = silver_path("idd", year)
    if not out.exists():
        return True
    xlsx = _find_xlsx(year)
    return xlsx.stat().st_mtime > out.stat().st_mtime


def process_year(year: int, verbose: bool = False, force: bool = False) -> None:
    """Processa um único ano do IDD."""
    xlsx_path = _find_xlsx(year)

    if not force and not _is_stale(year):
        if verbose:
            print(f"[{year}] silver já atualizado, pulando (use --force para reprocessar)")
        else:
            print(f"~ {year} (já atualizado)")
        return

    if verbose:
        print(f"[{year}] lendo {xlsx_path.name}")

    raw = _read_xlsx(xlsx_path)

    if verbose:
        print(f"[{year}] {raw.shape[0]} linhas, colunas: {raw.columns}")

    df = (
        raw
        .pipe(_rename_columns)
        .pipe(_normalize_sem_conceito)
        .pipe(_normalize_cebas)
        .pipe(_cast_schema)
        .pipe(_fix_outlier_continuo)
    )

    # Garante que 'ano' está preenchido (algumas planilhas têm o ano só na primeira linha)
    if "ano" in df.columns and df["ano"].null_count() > 0:
        first_ano = df["ano"].drop_nulls().first()
        df = df.with_columns(pl.col("ano").fill_null(first_ano))

    assert_not_empty(df, f"idd/{year}")
    assert_required_columns(df, IDD_REQUIRED_COLUMNS, f"idd/{year}")
    assert_no_nulls(df, ["co_ies", "co_curso", "ano"], f"idd/{year}")

    out = silver_path("idd", year)
    write_parquet(df, out)

    print(f"✓ {year} ({df.shape[0]} cursos)")
    if verbose:
        print(df.head(3))


@click.command()
@click.option("--year", required=True, help="Ano ou intervalo: 2023 ou 2016-2023")
@click.option("--force", is_flag=True, help="Reprocessa mesmo se silver já atualizado")
@click.option("--verbose", is_flag=True, help="Log detalhado")
def main(year: str, force: bool, verbose: bool) -> None:
    """Gera o silver do IDD."""
    from pipelines.censo.silver import _parse_years
    years = _parse_years(year)
    errors = []

    for y in years:
        try:
            process_year(y, verbose=verbose, force=force)
        except FileNotFoundError as e:
            print(f"✗ {y}: {e}", file=sys.stderr)
            errors.append(y)
        except Exception as e:
            print(f"✗ {y}: erro inesperado — {e}", file=sys.stderr)
            if verbose:
                import traceback
                traceback.print_exc()
            errors.append(y)

    if errors:
        print(f"\nFalhou em {len(errors)} ano(s): {errors}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
