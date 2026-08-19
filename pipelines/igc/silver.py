"""
pipelines/igc/silver.py — Silver do IGC (Índice Geral de Cursos).

Lê os XLSX do INEP e produz um Parquet silver por ano:
  silver/igc/{ano}.parquet

Uso:
  python -m pipelines.igc.silver --year 2023
  python -m pipelines.igc.silver --year 2017-2023
"""

import sys
from pathlib import Path

import click
import polars as pl

from pipelines.igc.schema import (
    IGC_COLUMN_MAP,
    IGC_REQUIRED_COLUMNS,
    IGC_SILVER_SCHEMA,
)
from shared.io import write_parquet
from shared.paths import bronze_dir, silver_path
from shared.validate import assert_no_nulls, assert_not_empty, assert_required_columns


# Sheets que não contêm dados do IGC
_SKIP_SHEETS = {
    "Atualizações", "Atualização",
    "CPC_2016_2017_2018", "CPC_2017_2018_2019", "CPC_2018_2019_2021",
    "Programas_Capes", "Programas_Capes_2018", "Programas_CAPES", "Programas CAPES",
    "matrículas_2019_2021_2022", "matriculas_2021_2022_2023", "matrículas_2021_2022_2023",
}


def _find_file(year: int) -> Path:
    """Localiza o arquivo Excel (.xlsx) do IGC para o ano dado."""
    base = bronze_dir("igc", year)
    if not base.exists():
        raise FileNotFoundError(
            f"Diretório bronze não encontrado: {base}\n"
            f"Execute: inex-ingest ingest --dataset igc --year {year}"
        )
    candidates = list(base.glob("*.xlsx")) + list(base.glob("*.XLSX"))
    if not candidates:
        raise FileNotFoundError(f"Nenhum arquivo .xlsx encontrado em {base}")
    return candidates[0]


def _read_xlsx(path: Path) -> pl.DataFrame:
    """Lê a sheet principal do IGC via openpyxl."""
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    sheet_name = next((s for s in wb.sheetnames if s not in _SKIP_SHEETS), wb.sheetnames[0])
    ws = wb[sheet_name]

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        raise ValueError(f"Planilha vazia: {path}")

    header = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
    data = rows[1:]

    str_data = {
        col: [
            str(row[i]).strip() if i < len(row) and row[i] is not None else None
            for row in data
        ]
        for i, col in enumerate(header)
    }

    df = pl.DataFrame(str_data)

    # Filtrar linhas onde co_ies é numérico (remove rodapés e linhas vazias)
    ies_col = next(
        (c for c in ["Código da IES", "Código da IES*"] if c in df.columns),
        None,
    )
    if ies_col:
        df = df.filter(pl.col(ies_col).str.contains(r"^\d+$"))

    return df.filter(pl.any_horizontal(pl.all().is_not_null()))


def _rename_columns(df: pl.DataFrame) -> pl.DataFrame:
    """Renomeia colunas usando o mapa canônico, descartando auxiliares (prefixo _)."""
    seen_targets: set[str] = set()
    rename_map: dict[str, str] = {}
    drop_cols: list[str] = []

    for orig_col in df.columns:
        target = IGC_COLUMN_MAP.get(orig_col)
        if target is None:
            continue
        if target.startswith("_"):
            drop_cols.append(orig_col)
        elif target in seen_targets:
            drop_cols.append(orig_col)
        else:
            rename_map[orig_col] = target
            seen_targets.add(target)

    df = df.drop([c for c in drop_cols if c in df.columns])
    return df.rename(rename_map)


def _cast_schema(df: pl.DataFrame) -> pl.DataFrame:
    """Aplica os tipos canônicos às colunas presentes."""
    casts = [
        pl.col(col).cast(dtype, strict=False)
        for col, dtype in IGC_SILVER_SCHEMA.items()
        if col in df.columns
    ]
    return df.with_columns(casts) if casts else df


def _normalize_cebas(df: pl.DataFrame) -> pl.DataFrame:
    """Converte in_cebas de texto ('-', 'Sim', etc.) para inteiro (0/1/null)."""
    if "in_cebas" not in df.columns:
        return df
    # Antes do cast: tratar valores textuais
    df = df.with_columns(
        pl.when(pl.col("in_cebas").is_in(["Sim", "SIM", "sim", "1"]))
        .then(pl.lit("1"))
        .when(pl.col("in_cebas").is_in(["-", "Não", "NÃO", "não", "0", ""]))
        .then(pl.lit("0"))
        .otherwise(pl.lit(None))
        .alias("in_cebas")
    )
    return df


def _is_stale(year: int) -> bool:
    """Retorna True se o silver precisa ser reprocessado."""
    out = silver_path("igc", year)
    if not out.exists():
        return True
    src = _find_file(year)
    return src.stat().st_mtime > out.stat().st_mtime


def process_year(year: int, verbose: bool = False, force: bool = False) -> None:
    """Processa um único ano do IGC."""
    file_path = _find_file(year)

    if not force and not _is_stale(year):
        if verbose:
            print(f"[{year}] silver já atualizado, pulando (use --force para reprocessar)")
        else:
            print(f"~ {year} (já atualizado)")
        return

    if verbose:
        print(f"[{year}] lendo {file_path.name}")

    raw = _read_xlsx(file_path)

    if verbose:
        print(f"[{year}] {raw.shape[0]} linhas, colunas: {raw.columns}")

    df = raw.pipe(_rename_columns).pipe(_normalize_cebas).pipe(_cast_schema)

    # 2017 não tem coluna 'Ano' no arquivo — injetar
    if "ano" not in df.columns:
        df = df.with_columns(pl.lit(year).cast(pl.Int32).alias("ano"))
    elif df["ano"].null_count() > 0:
        df = df.with_columns(pl.col("ano").fill_null(year))

    assert_not_empty(df, f"igc/{year}")
    assert_required_columns(df, IGC_REQUIRED_COLUMNS, f"igc/{year}")
    assert_no_nulls(df, ["co_ies", "ano"], f"igc/{year}")

    out = silver_path("igc", year)
    write_parquet(df, out)

    n_ies = df["co_ies"].n_unique()
    print(f"✓ {year} ({n_ies} IES)")
    if verbose:
        print(df.head(3))


def _parse_years(year_str: str) -> list[int]:
    """Converte '2017-2023' ou '2017,2019,2021-2023' em lista de ints."""
    years: list[int] = []
    for part in year_str.split(","):
        part = part.strip()
        if "-" in part:
            start, end = part.split("-", 1)
            years.extend(range(int(start), int(end) + 1))
        else:
            years.append(int(part))
    return sorted(set(years))


@click.command()
@click.option("--year", required=True, help="Ano ou intervalo: 2023 ou 2017-2023")
@click.option("--force", is_flag=True, help="Reprocessa mesmo se silver já atualizado")
@click.option("--verbose", is_flag=True, help="Log detalhado")
def main(year: str, force: bool, verbose: bool) -> None:
    """Gera o silver do IGC (suporta 2017–2023, exceto 2020)."""
    years = _parse_years(year)
    errors = []

    for y in years:
        if y == 2020:
            print(f"~ 2020 (não existe — ENADE não aplicado na pandemia)")
            continue
        try:
            process_year(y, verbose=verbose, force=force)
        except FileNotFoundError as e:
            print(f"✗ {y}: {e}", file=sys.stderr)
            errors.append(y)
        except Exception as e:
            print(f"✗ {y}: erro inesperado — {e}", file=sys.stderr)
            if verbose:
                import traceback
                traceback.print_exc()
            errors.append(y)

    if errors:
        print(f"\nFalhou em {len(errors)} ano(s): {errors}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
